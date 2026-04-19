from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\github\soap\Soap Calculator.xlsx")
OUTPUT_PATH = Path(r"C:\github\soap\soap-data.js")

HEADER_ROW = 6
FIRST_OIL_COLUMN = 4

METRICS = {
    "naohSap": 4,
    "kohSap": 5,
    "hardness": 7,
    "cleansing": 8,
    "condition": 9,
    "bubbly": 10,
    "creamy": 11,
    "iodine": 12,
    "ins": 13,
    "lauric": 15,
    "myristic": 16,
    "palmitic": 17,
    "stearic": 18,
    "ricinoleic": 19,
    "oleic": 20,
    "linoleic": 21,
    "linolenic": 22,
    "saturated": 24,
    "unsaturated": 25,
}


def normalize_number(value: object) -> float | None:
    if value in (None, "-"):
        return None
    return float(value)


def build_dataset() -> list[dict[str, object]]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    sheet = workbook["Soap Master"]
    oils: list[dict[str, object]] = []

    for column in range(FIRST_OIL_COLUMN, sheet.max_column + 1):
        name = sheet.cell(HEADER_ROW, column).value
        if not name:
            continue

        oil = {"name": str(name)}
        for key, row in METRICS.items():
            oil[key] = normalize_number(sheet.cell(row, column).value)
        oils.append(oil)

    return oils


def main() -> None:
    oils = build_dataset()
    output = "window.SOAP_OILS = " + json.dumps(oils, indent=2) + ";\n"
    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(f"Wrote {len(oils)} oils to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
